from io import BytesIO
from datetime import datetime
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Font

from backend.database import get_latest_scan


# =================================================
# EXCEL REPORT GENERATOR
# =================================================
def generate_excel_report() -> BytesIO:
    """
    Generate an Excel report from the latest scan.
    Returns BytesIO object for FastAPI Response.
    """

    scan = get_latest_scan()
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    timestamp = datetime.utcnow().isoformat()

    # ---------------- Layer 1 ----------------
    layer1 = scan.get("layer1", {})
    services = layer1.get("services", [])

    ws_services = wb.create_sheet("Services")
    _write_header(ws_services, [
        "Host", "Port", "Protocol", "State",
        "Service", "Product", "Version", "Vulnerabilities"
    ])

    for s in services:
        ws_services.append([
            s.get("host"),
            s.get("port"),
            s.get("protocol"),
            s.get("state"),
            s.get("service"),
            s.get("product"),
            s.get("version"),
            s.get("vulnerabilities", 0),
        ])

    # ---------------- Layer 2 ----------------
    layer2 = scan.get("layer2", {})
    intel = layer2.get("intel", [])

    ws_intel = wb.create_sheet("Threat_Intel")
    _write_header(ws_intel, [
        "Host", "Source", "Indicator", "Severity", "Details"
    ])

    for i in intel:
        ws_intel.append([
            i.get("host"),
            i.get("source"),
            i.get("indicator"),
            i.get("severity"),
            i.get("details"),
        ])

    # ---------------- Layer 3 ----------------
    layer3 = scan.get("layer3", {})
    risks = layer3.get("risk", [])

    ws_risk = wb.create_sheet("Risk_Scoring")
    _write_header(ws_risk, [
        "Host", "Risk Score", "Risk Level", "Threat Score"
    ])

    for r in risks:
        ws_risk.append([
            r.get("host"),
            r.get("risk_score"),
            r.get("risk_level"),
            r.get("threat_score"),
        ])

    # ---------------- Summary ----------------
    ws_summary = wb.create_sheet("Summary")
    _write_header(ws_summary, ["Metric", "Value"])

    ws_summary.append(["Scan Time", timestamp])
    ws_summary.append(["Total Hosts", len({s.get("host") for s in services})])
    ws_summary.append(["Total Services", len(services)])
    ws_summary.append(["Intel Records", len(intel)])
    ws_summary.append(["High Risk Hosts",
        sum(1 for r in risks if r.get("risk_level") in ("High", "Critical"))
    ])

    # ---------------- Save ----------------
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# =================================================
# HELPERS
# =================================================
def _write_header(ws, headers: List[str]) -> None:
    ws.append(headers)
    for col in ws[1]:
        col.font = Font(bold=True)
